# src/utils/exporter.py
import io
import pandas as pd
import logging
import os
from pathlib import Path

# Пытаемся импортировать openpyxl.styles, если не удается, используем заглушку
try:
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    logging.warning("openpyxl.styles не установлен, подсветка ячеек в Excel не будет работать")
    # Создаем заглушку для PatternFill
    class PatternFill:
        def __init__(self, start_color=None, end_color=None, fill_type=None):
            self.start_color = start_color
            self.end_color = end_color
            self.fill_type = fill_type
    OPENPYXL_AVAILABLE = False

# Словарь соответствия названий моделей и их локальных путей
CHRONOS_MODELS_MAPPING = {
    "bolt_tiny": "chronos-bolt-tiny",
    "bolt_small": "chronos-bolt-small", 
    "bolt_base": "chronos-bolt-base",
    "autogluon/chronos-bolt-tiny": "chronos-bolt-tiny",
    "autogluon/chronos-bolt-small": "chronos-bolt-small",
    "autogluon/chronos-bolt-base": "chronos-bolt-base"
}

def get_local_model_path(model_name):
    """
    Возвращает локальный путь к модели Chronos, если она доступна
    
    Аргументы:
        model_name (str): Имя модели
    
    Возвращает:
        str: Абсолютный путь к локальной модели или исходное имя модели
    """
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    chronos_dir = os.path.join(base_dir, "autogluon")
    
    if (model_name in CHRONOS_MODELS_MAPPING):
        model_dir = os.path.join(chronos_dir, CHRONOS_MODELS_MAPPING[model_name])
        if os.path.exists(model_dir):
            logging.info(f"ВАЖНО: Используется локальная модель Chronos из папки: {model_dir}")
            # Проверяем наличие модельного файла
            model_file = os.path.join(model_dir, "model.safetensors")
            if os.path.exists(model_file):
                model_size_mb = round(os.path.getsize(model_file) / (1024 * 1024), 2)
                logging.info(f"Размер файла модели: {model_size_mb} МБ")
            return model_dir
    
    logging.info(f"Локальная модель не найдена для {model_name}, будет использован Hugging Face")
    return model_name

def generate_excel_buffer(result):
    """
    Формирует Excel-файл в памяти с результатами прогнозирования или обучения
    
    Аргументы:
        result (dict): Результат прогнозирования/обучения, содержащий различные данные:
            - forecasts: словарь прогнозов для разных целевых переменных
            - leaderboard: таблица с результатами обучения моделей
            - ensemble_info: информация о весах моделей в ансамбле
            - module_error: информация об ошибках модулей
            - summary_data: сводная информация о результатах
    
    Возвращает:
        BytesIO: Объект BytesIO с Excel-файлом или None в случае ошибки
    """
    # Инициализируем буфер для Excel
    excel_buffer = io.BytesIO()
    
    # Флаг успешной записи в Excel
    excel_written = False
    
    try:
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # Проверяем успешность результата
            if not result.get('success', False):
                # Создаем лист с ошибкой
                pd.DataFrame([{"Ошибка": result.get('error', 'Неизвестная ошибка')}]).to_excel(
                    writer, sheet_name="Ошибка", index=False)
                excel_written = True
                return excel_buffer
            
            # Счетчик созданных листов
            sheets_created = 0
            
            # ===== 1. Обработка прогнозов =====
            if 'forecasts' in result and isinstance(result['forecasts'], dict):
                sheet_idx = 0  # Индекс для именования листов при длинных именах
                
                for target_col, forecast_data in result['forecasts'].items():
                    predictions = forecast_data.get('predictions')
                    if predictions is not None and not predictions.empty:
                        sheet_name = f"Прогноз_{target_col}"
                        if len(sheet_name) > 31:  # Ограничение Excel на длину имени листа
                            sheet_name = f"Прогноз_{sheet_idx}"
                            sheet_idx += 1
                        
                        # Сбрасываем индекс для экспорта в Excel
                        predictions.reset_index().to_excel(writer, sheet_name=sheet_name, index=False)
                        sheets_created += 1
                        excel_written = True
                        logging.info(f"Лист с прогнозом для {target_col} успешно создан")
            
            # ===== 2. Обработка сводных данных =====
            if 'summary_data' in result and isinstance(result['summary_data'], dict):
                # Преобразуем сводные данные в DataFrame для Excel
                summary_data = []
                
                for metric_key, metric_value in result['summary_data'].items():
                    if isinstance(metric_value, dict):
                        for sub_key, sub_value in metric_value.items():
                            summary_data.append({"Метрика": f"{metric_key}: {sub_key}", "Значение": sub_value})
                    else:
                        summary_data.append({"Метрика": metric_key, "Значение": metric_value})
                
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name="Сводка", index=False)
                    sheets_created += 1
                    excel_written = True
                    logging.info(f"Лист со сводными данными успешно создан")
            
            # ===== 3. Обработка информации об ошибках модулей =====
            if 'module_error' in result:
                error_msg = result['module_error']
                if isinstance(error_msg, str):
                    pd.DataFrame([{"Сообщение": error_msg}]).to_excel(writer, sheet_name="Ошибки_модулей", index=False)
                    sheets_created += 1
                    excel_written = True
                    logging.info(f"Лист с информацией об ошибках модулей успешно создан")
            
            # ===== 4. Обработка лидерборда =====
            if 'leaderboard' in result and isinstance(result.get('leaderboard'), pd.DataFrame):
                leaderboard = result['leaderboard']
                if not leaderboard.empty:
                    # Округляем числовые значения для лучшего отображения
                    for col in leaderboard.select_dtypes(include=['float']).columns:
                        leaderboard[col] = leaderboard[col].round(6)
                    
                    leaderboard.to_excel(writer, sheet_name="Лидерборд", index=False)
                    sheets_created += 1
                    excel_written = True
                    logging.info(f"Лист с лидербордом успешно создан")
            
            # ===== 5. Обработка ансамблевых весов =====
            ensemble_weights_df = None
            model_details_df = None
            
            # Способ 1: Уже извлеченные веса ансамбля
            if 'ensemble_info' in result:
                ensemble_info = result['ensemble_info']
                if isinstance(ensemble_info, tuple) and len(ensemble_info) == 2:
                    ensemble_weights_df, model_details_df = ensemble_info
                elif isinstance(ensemble_info, dict):
                    if 'weights' in ensemble_info:
                        ensemble_weights_df = ensemble_info['weights']
                    if 'details' in ensemble_info:
                        model_details_df = ensemble_info['details']
            
            # Способ 2: Извлечение весов из предиктора, если доступен
            if ensemble_weights_df is None and 'predictor' in result:
                try:
                    ensemble_weights_df, model_details_df = extract_ensemble_weights(result['predictor'])
                except Exception as e:
                    logging.warning(f"Не удалось извлечь информацию об ансамбле из предиктора: {e}")
            
            # Сохраняем веса ансамбля, если они есть
            if ensemble_weights_df is not None and isinstance(ensemble_weights_df, pd.DataFrame) and not ensemble_weights_df.empty:
                ensemble_weights_df.to_excel(writer, sheet_name="Веса_ансамбля", index=False)
                sheets_created += 1
                excel_written = True
                logging.info(f"Лист с весами ансамбля успешно создан")
            
            # Сохраняем детали моделей, если они есть
            if model_details_df is not None and isinstance(model_details_df, pd.DataFrame) and not model_details_df.empty:
                model_details_df.to_excel(writer, sheet_name="Детали_моделей", index=False)
                sheets_created += 1
                excel_written = True
                logging.info(f"Лист с деталями моделей успешно создан")
            
            # Если не создано ни одного листа, добавляем информационный лист
            if sheets_created == 0:
                pd.DataFrame([{"Информация": "Нет данных для отображения в Excel"}]).to_excel(
                    writer, sheet_name="Информация", index=False)
                excel_written = True
                logging.warning(f"В результатах не найдено данных для экспорта в Excel")
            
            # Дополнительное логирование
            logging.info(f"Excel-файл успешно создан, листов: {sheets_created}")
    
    except Exception as e:
        error_msg = f"Ошибка при создании Excel-файла: {str(e)}"
        logging.error(error_msg)
        
        # Пытаемся создать простой Excel-файл с информацией об ошибке
        try:
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                pd.DataFrame([{"Ошибка": error_msg}]).to_excel(writer, sheet_name="Ошибка", index=False)
                excel_written = True
        except Exception as inner_e:
            logging.error(f"Не удалось создать Excel-файл даже с информацией об ошибке: {str(inner_e)}")
            return None
    
    # Проверяем, что хотя бы один лист был успешно создан
    if not excel_written:
        logging.error("Excel-файл не был создан из-за отсутствия данных или ошибок")
        return None
    
    # Возвращаем указатель в начало буфера
    excel_buffer.seek(0)
    return excel_buffer

def extract_ensemble_weights(predictor):
    """
    Извлекает информацию о составе и весах ансамблевой модели
    
    Аргументы:
        predictor: Обученный TimeSeriesPredictor
    
    Возвращает:
        tuple: (weights_df, details_df) или (None, None)
            weights_df: pandas.DataFrame с весами моделей
            details_df: pandas.DataFrame с дополнительной информацией о моделях
    """
    try:
        if not predictor:
            logging.warning("Предиктор не предоставлен для извлечения весов ансамбля")
            return None, None
            
        # Проверяем, является ли лучшая модель ансамблевой
        try:
            # Используем новое property model_best вместо устаревшего метода get_model_best
            best_model_name = predictor.model_best if hasattr(predictor, 'model_best') else predictor.model_names()[0]
            logging.info(f"Лучшая модель: {best_model_name}")
            
            # Проверка, является ли модель ансамблевой
            is_ensemble = "Ensemble" in best_model_name or "WeightedEnsemble" in best_model_name
            
            if not is_ensemble:
                logging.info(f"Лучшая модель не является ансамблевой: {best_model_name}")
                return None, None
        except Exception as e:
            logging.warning(f"Не удалось определить, является ли лучшая модель ансамблевой: {e}")
            # Продолжаем выполнение, так как ансамбль все равно может существовать
        
        # Инициализируем словарь для хранения весов
        ensemble_weights = {}
        model_details = {}
        
        # Метод 1: Использование метода get_fitted_model или get_model
        try:
            # В новых версиях AutoGluon используется только model_names()
            ensemble_model_name = None
            
            # Получаем список моделей
            if hasattr(predictor, 'model_names'):
                model_names = predictor.model_names()
            else:
                # Обратная совместимость
                model_names = predictor.get_model_names()
            
            # Ищем ансамблевую модель
            for model_name in model_names:
                if "Ensemble" in model_name or "WeightedEnsemble" in model_name:
                    ensemble_model_name = model_name
                    break
            
            if ensemble_model_name:
                try:
                    # В новых версиях AutoGluon
                    ensemble_model = predictor.get_model(ensemble_model_name)
                except:
                    try:
                        # В старых версиях AutoGluon
                        ensemble_model = predictor.get_fitted_model(ensemble_model_name)
                    except:
                        ensemble_model = None
                
                if ensemble_model:
                    # Извлекаем веса из модели
                    if hasattr(ensemble_model, 'weights_'):
                        weights = ensemble_model.weights_
                        for model_name, weight in weights.items():
                            ensemble_weights[model_name] = weight
                        logging.info(f"Извлечены веса из ensemble_model.weights_: {len(ensemble_weights)} моделей")
                    
                    # Альтернативный способ для более новых версий
                    elif hasattr(ensemble_model, 'model_base_ensemble') and hasattr(ensemble_model.model_base_ensemble, 'weights'):
                        weights = ensemble_model.model_base_ensemble.weights
                        for model_name, weight in weights.items():
                            ensemble_weights[model_name] = weight
                        logging.info(f"Извлечены веса из model_base_ensemble.weights: {len(ensemble_weights)} моделей")
                    
                    # Еще один альтернативный способ
                    elif hasattr(ensemble_model, 'model') and hasattr(ensemble_model.model, 'weights'):
                        weights = ensemble_model.model.weights
                        for model_name, weight in weights.items():
                            ensemble_weights[model_name] = weight
                        logging.info(f"Извлечены веса из ensemble_model.model.weights: {len(ensemble_weights)} моделей")
        except Exception as e:
            logging.warning(f"Не удалось извлечь веса ансамбля методом 1: {e}")
        
        # Метод 2: Использование информации из лидерборда
        if not ensemble_weights:
            try:
                # Получаем лидерборд
                leaderboard = predictor.leaderboard() if hasattr(predictor, 'leaderboard') else None
                
                if leaderboard is not None and not leaderboard.empty:
                    # Ищем ансамблевые модели в лидерборде
                    ensemble_models = leaderboard[leaderboard['model'].str.contains('Ensemble|WeightedEnsemble', case=False)]
                    
                    if not ensemble_models.empty:
                        best_ensemble = ensemble_models.iloc[0]
                        ensemble_model_name = best_ensemble['model']
                        
                        # Теперь пытаемся извлечь информацию о составе ансамбля
                        try:
                            # В новых версиях AutoGluon
                            ensemble_model = predictor.get_model(ensemble_model_name)
                        except:
                            try:
                                # В старых версиях AutoGluon
                                ensemble_model = predictor.get_fitted_model(ensemble_model_name)
                            except:
                                ensemble_model = None
                        
                        if ensemble_model:
                            # Попытка извлечь веса разными способами
                            if hasattr(ensemble_model, 'weights_'):
                                for model_name, weight in ensemble_model.weights_.items():
                                    ensemble_weights[model_name] = weight
                            elif hasattr(ensemble_model, 'model_weights'):
                                for model_name, weight in ensemble_model.model_weights.items():
                                    ensemble_weights[model_name] = weight
                            elif hasattr(ensemble_model, 'get_model_weights'):
                                weights = ensemble_model.get_model_weights()
                                for model_name, weight in weights.items():
                                    ensemble_weights[model_name] = weight
            except Exception as e:
                logging.warning(f"Не удалось извлечь веса ансамбля методом 2: {e}")
        
        # Метод 3: Прямой доступ к информации о моделях
        if not ensemble_weights:
            try:
                # В некоторых версиях AutoGluon информация о моделях хранится напрямую
                if hasattr(predictor, '_trainer') and hasattr(predictor._trainer, 'model_weights'):
                    for model_name, weight in predictor._trainer.model_weights.items():
                        ensemble_weights[model_name] = weight
                    logging.info(f"Извлечены веса из _trainer.model_weights: {len(ensemble_weights)} моделей")
            except Exception as e:
                logging.warning(f"Не удалось извлечь веса ансамбля методом 3: {e}")
        
        # Если не удалось извлечь веса, возвращаем None
        if not ensemble_weights:
            logging.warning("Не удалось извлечь веса ансамбля ни одним из методов")
            return None, None
        
        # Создаем DataFrame с весами моделей
        weights_df = pd.DataFrame({
            "Модель": list(ensemble_weights.keys()),
            "Вес": list(ensemble_weights.values())
        })
        
        # Добавляем процентное представление весов
        weights_df["Вес (%)"] = weights_df["Вес"] * 100
        
        # Сортируем по весу в порядке убывания
        weights_df = weights_df.sort_values("Вес", ascending=False)
        
        # Получаем информацию о моделях для деталей
        try:
            model_info = {}
            for model_name in weights_df["Модель"]:
                try:
                    # Пытаемся получить модель
                    model = predictor.get_model(model_name) if hasattr(predictor, 'get_model') else None
                    
                    if model:
                        # Собираем информацию о модели
                        model_info[model_name] = {
                            "Тип": type(model).__name__,
                            "Параметры": str(getattr(model, 'params', 'Н/Д'))[:200]  # Ограничиваем длину строки
                        }
                except Exception as model_err:
                    logging.warning(f"Не удалось получить информацию о модели {model_name}: {model_err}")
            
            # Создаем DataFrame с деталями моделей
            if model_info:
                models_data = []
                for model_name, info in model_info.items():
                    row = {"Модель": model_name}
                    row.update(info)
                    models_data.append(row)
                
                details_df = pd.DataFrame(models_data)
            else:
                details_df = None
        except Exception as detail_err:
            logging.warning(f"Не удалось создать DataFrame с деталями моделей: {detail_err}")
            details_df = None
        
        return weights_df, details_df
    except Exception as e:
        logging.error(f"Ошибка при извлечении весов ансамбля: {e}")
        return None, None

# Старая функция оставлена для совместимости, но помечена как устаревшая
def generate_excel_buffer_legacy(preds, leaderboard, static_train, ensemble_info_df):
    """
    УСТАРЕВШАЯ ФУНКЦИЯ - используйте generate_excel_buffer(result)
    
    Формирует Excel-файл в памяти с листами:
      - Predictions
      - Leaderboard с подсветкой лучшей модели
      - StaticTrainFeatures
      - WeightedEnsembleInfo
    
    Возвращает объект BytesIO с Excel-файлом.
    """
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        # Лист с предсказаниями
        if preds is not None:
            preds.reset_index().to_excel(writer, sheet_name="Predictions", index=False)
        # Лидерборд с подсветкой лучшей модели
        if leaderboard is not None:
            leaderboard.to_excel(writer, sheet_name="Leaderboard", index=False)
            try:
                if OPENPYXL_AVAILABLE:
                    sheet_lb = writer.sheets["Leaderboard"]
                    best_idx = leaderboard.iloc[0].name  # индекс строки лучшей модели
                    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    row_excel = best_idx + 2  # +2 из-за заголовка
                    for col_idx in range(1, leaderboard.shape[1] + 1):
                        cell = sheet_lb.cell(row=row_excel, column=col_idx)
                        cell.fill = fill_green
            except Exception as e:
                print(f"Ошибка при подсветке лучшей модели в Leaderboard: {e}")
        # Лист со статическими признаками
        if static_train is not None and not static_train.empty:
            static_train.to_excel(writer, sheet_name="StaticTrainFeatures", index=False)
        # Лист с информацией об ансамбле
        if ensemble_info_df is not None and not ensemble_info_df.empty:
            ensemble_info_df.to_excel(writer, sheet_name="WeightedEnsembleInfo", index=False)
    return excel_buffer
