"""
Модуль для экспорта результатов прогнозирования в Excel.
"""

import io
import logging
import numpy as np
import pandas as pd
from typing import Dict, Any, List, Union, Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference


def export_forecast_to_excel(
    forecast_results: Dict[str, Any],
    include_leaderboard: bool = True,
    include_model_info: bool = True,
    include_individual_models: bool = True,
    include_graphs: bool = True
) -> io.BytesIO:
    """
    Экспортирует результаты прогнозирования в Excel-файл.
    
    Parameters:
    -----------
    forecast_results : Dict[str, Any]
        Словарь с результатами прогнозирования
    include_leaderboard : bool, default=True
        Включать ли таблицу лидеров в отчет
    include_model_info : bool, default=True
        Включать ли информацию о модели в отчет
    include_individual_models : bool, default=True
        Включать ли информацию об отдельных моделях в отчет
    include_graphs : bool, default=True
        Включать ли графики в отчет
        
    Returns:
    --------
    io.BytesIO
        Буфер с Excel-файлом
    """
    logging.info("Начинаем экспорт результатов прогнозирования в Excel")
    
    # Создаем буфер для файла
    output = io.BytesIO()
    
    try:
        # Создаем книгу Excel и объект записи
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Получаем книгу
            workbook = writer.book
            
            # Применяем стили для заголовков
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Добавляем лист с параметрами прогноза
            _add_parameters_sheet(forecast_results, workbook, writer, header_fill, header_font, header_alignment)
            
            # Добавляем листы с прогнозами для каждой целевой переменной
            _add_forecasts_sheets(forecast_results, workbook, writer, header_fill, header_font, header_alignment)
            
            # Добавляем лист с информацией о моделях
            if include_model_info and 'model_metrics' in forecast_results:
                _add_model_info_sheet(forecast_results, workbook, writer, header_fill, header_font, header_alignment)
            
            # Добавляем лист с таблицей лидеров
            if include_leaderboard and 'model_metrics' in forecast_results:
                _add_leaderboard_sheet(forecast_results, workbook, writer, header_fill, header_font, header_alignment)
            
            # Добавляем лист с информацией об отдельных моделях
            if include_individual_models and 'model_metrics' in forecast_results:
                _add_individual_models_sheet(forecast_results, workbook, writer, header_fill, header_font, header_alignment)
            
            # Добавляем лист с информацией о дрейфе концепции
            _add_drift_info_sheet(forecast_results, workbook, writer, header_fill, header_font, header_alignment)
            
        # Сбрасываем указатель буфера на начало
        output.seek(0)
        
        logging.info("Экспорт результатов прогнозирования в Excel успешно завершен")
        return output
        
    except Exception as e:
        logging.error(f"Ошибка при экспорте результатов прогнозирования в Excel: {e}")
        # Если возникла ошибка, возвращаем пустой буфер
        output = io.BytesIO()
        output.seek(0)
        return output


def _add_parameters_sheet(
    forecast_results: Dict[str, Any],
    workbook: Workbook,
    writer: pd.ExcelWriter,
    header_fill: PatternFill,
    header_font: Font,
    header_alignment: Alignment
) -> None:
    """Добавляет лист с параметрами прогноза"""
    # Создаем датафрейм с параметрами
    params = forecast_results.get('parameters', {})
    if not params:
        return
    
    # Преобразуем параметры в датафрейм для записи
    params_list = []
    for key, value in params.items():
        # Обрабатываем списки и другие сложные типы данных
        if isinstance(value, (list, tuple)):
            value = ', '.join(str(v) for v in value)
        
        params_list.append({
            'Параметр': key,
            'Значение': value
        })
    
    params_df = pd.DataFrame(params_list)
    
    # Записываем датафрейм на лист
    params_df.to_excel(writer, sheet_name='Параметры прогноза', index=False)
    
    # Применяем форматирование
    worksheet = writer.sheets['Параметры прогноза']
    for col_num, column in enumerate(params_df.columns, 1):
        column_letter = get_column_letter(col_num)
        cell = worksheet[f"{column_letter}1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Автоматически подгоняем ширину колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _add_forecasts_sheets(
    forecast_results: Dict[str, Any],
    workbook: Workbook,
    writer: pd.ExcelWriter,
    header_fill: PatternFill,
    header_font: Font,
    header_alignment: Alignment
) -> None:
    """Добавляет листы с прогнозами для каждой целевой переменной"""
    forecasts = forecast_results.get('forecasts', {})
    if not forecasts:
        return
    
    for target_col, forecast_data in forecasts.items():
        # Если есть ошибка в прогнозе для этой колонки, пропускаем
        if 'error' in forecast_data:
            continue
        
        # Получаем данные прогноза
        predictions = forecast_data.get('predictions', None)
        if predictions is None:
            continue
        
        # Преобразуем результаты прогноза в DataFrame
        try:
            # Преобразуем результаты прогноза в DataFrame для экспорта
            if isinstance(predictions, pd.DataFrame):
                forecast_df = predictions.copy()
                
                # Сбрасываем индекс, чтобы дата и ID стали колонками
                if isinstance(forecast_df.index, pd.MultiIndex):
                    forecast_df = forecast_df.reset_index()
                else:
                    forecast_df = forecast_df.reset_index()
                
                # Переименовываем колонки для лучшей читаемости
                col_map = {
                    'item_id': 'ID',
                    'timestamp': 'Дата'
                }
                
                forecast_df.rename(columns=col_map, inplace=True)
                
                # Записываем прогноз на лист
                sheet_name = f'Прогноз_{target_col}'
                if len(sheet_name) > 31:  # Excel ограничивает длину имени листа до 31 символа
                    sheet_name = f'Прогноз_{len(target_col)}'
                
                forecast_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Применяем форматирование к заголовкам
                worksheet = writer.sheets[sheet_name]
                for col_num, column in enumerate(forecast_df.columns, 1):
                    column_letter = get_column_letter(col_num)
                    cell = worksheet[f"{column_letter}1"]
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Автоматически подгоняем ширину колонок
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
            else:
                logging.warning(f"Прогнозы для {target_col} не являются DataFrame")
                
        except Exception as e:
            logging.error(f"Ошибка при добавлении листа с прогнозами для {target_col}: {e}")


def _add_model_info_sheet(
    forecast_results: Dict[str, Any],
    workbook: Workbook,
    writer: pd.ExcelWriter,
    header_fill: PatternFill,
    header_font: Font,
    header_alignment: Alignment
) -> None:
    """Добавляет лист с общей информацией о моделях"""
    model_metrics = forecast_results.get('model_metrics', {})
    if not model_metrics:
        return
    
    # Создаем датафрейм с информацией о моделях
    model_info_rows = []
    
    for target_col, metrics in model_metrics.items():
        # Пропускаем, если есть ошибка
        if 'error' in metrics:
            continue
        
        # Получаем базовую информацию о модели
        model_info = metrics.get('model_info', {})
        if not model_info:
            continue
        
        # Базовая информация
        row = {
            'Целевая переменная': target_col,
            'Длина прогноза': model_info.get('prediction_length', 'N/A'),
            'Частота': model_info.get('freq', 'N/A'),
            'Метрика оценки': model_info.get('eval_metric', 'N/A'),
            'Квантили': ', '.join(str(q) for q in model_info.get('quantile_levels', [])) if 'quantile_levels' in model_info else 'N/A',
            'Количество моделей': model_info.get('model_count', 'N/A')
        }
        
        # Добавляем информацию о лучшей модели, если есть
        if 'best_model' in metrics:
            row['Лучшая модель'] = metrics['best_model']
            row['Лучший результат'] = metrics.get('best_score', 'N/A')
        
        # Добавляем список моделей
        if 'models' in model_info:
            row['Модели'] = ', '.join(model_info['models'])
        
        model_info_rows.append(row)
    
    if not model_info_rows:
        return
    
    # Создаем датафрейм
    model_info_df = pd.DataFrame(model_info_rows)
    
    # Записываем на лист
    model_info_df.to_excel(writer, sheet_name='Информация о моделях', index=False)
    
    # Применяем форматирование
    worksheet = writer.sheets['Информация о моделях']
    for col_num, column in enumerate(model_info_df.columns, 1):
        column_letter = get_column_letter(col_num)
        cell = worksheet[f"{column_letter}1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Автоматически подгоняем ширину колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _add_leaderboard_sheet(
    forecast_results: Dict[str, Any],
    workbook: Workbook,
    writer: pd.ExcelWriter,
    header_fill: PatternFill,
    header_font: Font,
    header_alignment: Alignment
) -> None:
    """Добавляет лист с таблицей лидеров моделей"""
    model_metrics = forecast_results.get('model_metrics', {})
    if not model_metrics:
        return
    
    # Создаем словарь с таблицами лидеров для каждой целевой переменной
    leaderboards = {}
    
    for target_col, metrics in model_metrics.items():
        # Пропускаем, если есть ошибка или нет таблицы лидеров
        if 'error' in metrics or 'leaderboard' not in metrics:
            continue
        
        leaderboard_dict = metrics['leaderboard']
        if not leaderboard_dict:
            continue
        
        # Преобразуем словарь в DataFrame
        rows = []
        for model_name, scores in leaderboard_dict.items():
            row = {'model': model_name}
            row.update(scores)
            rows.append(row)
        
        if rows:
            leaderboards[target_col] = pd.DataFrame(rows)
    
    if not leaderboards:
        return
    
    # Добавляем таблицы лидеров на отдельные листы или на один лист
    for target_col, leaderboard_df in leaderboards.items():
        sheet_name = f'Лидерборд_{target_col}'
        if len(sheet_name) > 31:  # Excel ограничивает длину имени листа до 31 символа
            sheet_name = f'Лидерборд_{len(target_col)}'
        
        leaderboard_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Применяем форматирование
        worksheet = writer.sheets[sheet_name]
        for col_num, column in enumerate(leaderboard_df.columns, 1):
            column_letter = get_column_letter(col_num)
            cell = worksheet[f"{column_letter}1"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Автоматически подгоняем ширину колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width


def _add_individual_models_sheet(
    forecast_results: Dict[str, Any],
    workbook: Workbook,
    writer: pd.ExcelWriter,
    header_fill: PatternFill,
    header_font: Font,
    header_alignment: Alignment
) -> None:
    """Добавляет лист с информацией об отдельных моделях и их гиперпараметрах"""
    model_metrics = forecast_results.get('model_metrics', {})
    if not model_metrics:
        return
    
    for target_col, metrics in model_metrics.items():
        # Пропускаем, если есть ошибка или нет информации о моделях
        if 'error' in metrics or 'models_hyperparameters' not in metrics:
            continue
        
        model_hyperparams = metrics['models_hyperparameters']
        if not model_hyperparams:
            continue
        
        # Создаем DataFrame с гиперпараметрами
        rows = []
        
        for model_name, params in model_hyperparams.items():
            # Базовая информация о модели
            row = {'Модель': model_name}
            
            # Добавляем все гиперпараметры модели
            if isinstance(params, dict):
                for param_name, param_value in params.items():
                    # Преобразуем сложные типы данных в строку
                    if isinstance(param_value, (dict, list, tuple)):
                        param_value = str(param_value)
                    row[param_name] = param_value
            
            rows.append(row)
        
        if not rows:
            continue
        
        # Создаем DataFrame
        hyperparams_df = pd.DataFrame(rows)
        
        # Записываем на лист
        sheet_name = f'Параметры_моделей_{target_col}'
        if len(sheet_name) > 31:  # Excel ограничивает длину имени листа
            sheet_name = f'Параметры_моделей_{len(target_col)}'
        
        hyperparams_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Применяем форматирование
        worksheet = writer.sheets[sheet_name]
        for col_num, column in enumerate(hyperparams_df.columns, 1):
            column_letter = get_column_letter(col_num)
            cell = worksheet[f"{column_letter}1"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Автоматически подгоняем ширину колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min((max_length + 2) * 1.2, 50)  # Ограничиваем максимальную ширину
            worksheet.column_dimensions[column_letter].width = adjusted_width


def _add_drift_info_sheet(
    forecast_results: Dict[str, Any],
    workbook: Workbook,
    writer: pd.ExcelWriter,
    header_fill: PatternFill,
    header_font: Font,
    header_alignment: Alignment
) -> None:
    """Добавляет лист с информацией о дрейфе концепции"""
    forecasts = forecast_results.get('forecasts', {})
    if not forecasts:
        return
    
    # Собираем информацию о дрейфе для всех целевых переменных
    drift_rows = []
    
    for target_col, forecast_data in forecasts.items():
        # Пропускаем, если есть ошибка
        if 'error' in forecast_data:
            continue
        
        # Получаем информацию о дрейфе
        drift_detected = forecast_data.get('drift_detected', False)
        drift_info = forecast_data.get('drift_info', {})
        
        # Если нет информации о дрейфе, пропускаем
        if not drift_info:
            continue
        
        # Для каждого ID с информацией о дрейфе создаём строку в таблице
        for item_id, item_drift in drift_info.items():
            # Проверяем, что это словарь и в нём есть информация о дрейфе
            if not isinstance(item_drift, dict):
                continue
            
            item_drift_detected = item_drift.get('drift_detected', False)
            
            # Создаем строку с базовой информацией
            row = {
                'Целевая переменная': target_col,
                'ID': item_id,
                'Дрейф обнаружен': 'Да' if item_drift_detected else 'Нет'
            }
            
            # Добавляем статистики по историческим данным
            hist_stats = item_drift.get('historical_stats', {})
            if hist_stats:
                for key, value in hist_stats.items():
                    row[f'Исторические данные - {key}'] = value
            
            # Добавляем статистики по прогнозным данным
            forecast_stats = item_drift.get('forecast_stats', {})
            if forecast_stats:
                for key, value in forecast_stats.items():
                    row[f'Прогноз - {key}'] = value
            
            # Добавляем информацию о причинах дрейфа
            drift_reasons = item_drift.get('drift_reasons', [])
            if drift_reasons:
                # Обрабатываем разные форматы причин дрейфа
                reasons_str = []
                for reason in drift_reasons:
                    if isinstance(reason, dict):
                        reason_text = reason.get('reason', '')
                        details = reason.get('details', '')
                        if details:
                            reason_text += f": {details}"
                        reasons_str.append(reason_text)
                    else:
                        reasons_str.append(str(reason))
                
                row['Причины дрейфа'] = '; '.join(reasons_str)
            
            drift_rows.append(row)
    
    if not drift_rows:
        return
    
    # Создаем DataFrame
    drift_df = pd.DataFrame(drift_rows)
    
    # Упорядочиваем колонки в логичном порядке
    ordered_columns = []
    
    # Базовые колонки
    base_columns = ['Целевая переменная', 'ID', 'Дрейф обнаружен', 'Причины дрейфа']
    for col in base_columns:
        if col in drift_df.columns:
            ordered_columns.append(col)
    
    # Статистики исторических данных
    hist_columns = [col for col in drift_df.columns if col.startswith('Исторические данные')]
    hist_columns.sort()
    ordered_columns.extend(hist_columns)
    
    # Статистики прогнозных данных
    forecast_columns = [col for col in drift_df.columns if col.startswith('Прогноз')]
    forecast_columns.sort()
    ordered_columns.extend(forecast_columns)
    
    # Добавляем оставшиеся колонки
    for col in drift_df.columns:
        if col not in ordered_columns:
            ordered_columns.append(col)
    
    # Переупорядочиваем колонки
    drift_df = drift_df[ordered_columns]
    
    # Записываем на лист
    drift_df.to_excel(writer, sheet_name='Информация о дрейфе', index=False)
    
    # Применяем форматирование
    worksheet = writer.sheets['Информация о дрейфе']
    for col_num, column in enumerate(drift_df.columns, 1):
        column_letter = get_column_letter(col_num)
        cell = worksheet[f"{column_letter}1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Автоматически подгоняем ширину колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min((max_length + 2) * 1.2, 50)  # Ограничиваем максимальную ширину
        worksheet.column_dimensions[column_letter].width = adjusted_width


def generate_excel_report(
    forecast_results: Dict[str, Any],
    include_leaderboard: bool = True,
    include_model_info: bool = True,
    include_individual_models: bool = True,
    include_graphs: bool = False
) -> io.BytesIO:
    """
    Генерирует Excel-отчет с результатами прогнозирования.
    
    Parameters:
    -----------
    forecast_results : Dict[str, Any]
        Словарь с результатами прогнозирования
    include_leaderboard : bool, default=True
        Включать ли таблицу лидеров в отчет
    include_model_info : bool, default=True
        Включать ли информацию о модели в отчет
    include_individual_models : bool, default=True
        Включать ли информацию об отдельных моделях в отчет
    include_graphs : bool, default=False
        Включать ли графики в отчет (экспериментальная функция)
        
    Returns:
    --------
    io.BytesIO
        Буфер с Excel-файлом
    """
    logging.info("Генерация Excel-отчета с результатами прогнозирования")
    
    # Внутренние параметры в случае, если forecast_results не содержит некоторых ключей
    if 'parameters' not in forecast_results and 'metadata' in forecast_results:
        forecast_results['parameters'] = forecast_results['metadata']
    
    # Добавляем информацию о времени генерации отчета
    from datetime import datetime
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if 'parameters' not in forecast_results:
        forecast_results['parameters'] = {}
    forecast_results['parameters']['report_generated'] = current_time
    
    # Проверяем наличие ключевых данных
    if 'forecasts' not in forecast_results:
        logging.warning("В результатах прогнозирования отсутствуют прогнозы")
        forecast_results['forecasts'] = {}
    
    # Экспортируем результаты в Excel
    return export_forecast_to_excel(
        forecast_results=forecast_results,
        include_leaderboard=include_leaderboard,
        include_model_info=include_model_info,
        include_individual_models=include_individual_models,
        include_graphs=include_graphs
    )
