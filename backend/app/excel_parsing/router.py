from fastapi import APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse
import pandas as pd
import os
from fastapi import Query
from pydantic import BaseModel
import numpy as np
from concurrent.futures import ThreadPoolExecutor
import asyncio
from fastapi import Request
from io import BytesIO
from openpyxl import load_workbook



router = APIRouter()

@router.post("/preview-excel")
async def preview_excel(file: UploadFile = File(...)):
    try:
        filename = file.filename.lower()
        file.file.seek(0)
        if filename.endswith('.csv'):
            df = pd.read_csv(file.file, nrows=10)
            file.file.seek(0)
            # Для csv считаем строки вручную
            total_rows = sum(1 for _ in file.file) - 1  # минус строка заголовков
        else:
            # Читаем файл в память
            file_bytes = await file.read()
            excel_io = BytesIO(file_bytes)
            df = pd.read_excel(excel_io, nrows=10)
            excel_io.seek(0)
            wb = load_workbook(excel_io, read_only=True)
            ws = wb.active
            total_rows = ws.max_row - 1  # минус строка заголовков
            wb.close()
        df = df.astype(str)
        data = df.to_dict(orient="records")
        columns = list(df.columns)
        return JSONResponse({"columns": columns, "rows": data, "total_rows": total_rows})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка обработки файла: {str(e)}")

class TimeSeriesRequest(BaseModel):
    session_id: str
    id_column: str
    ts_id: str

@router.post("/get-timeseries-by-id")
async def get_timeseries_by_id(req: TimeSeriesRequest):
    session_id = req.session_id
    id_column = req.id_column
    ts_id = req.ts_id
    if not session_id or not id_column or not ts_id:
        raise HTTPException(status_code=400, detail="session_id, id_column и ts_id обязательны")
    session_dir = os.path.join("training_sessions", session_id)
    parquet_path = os.path.join(session_dir, "training_data.parquet")
    if not os.path.exists(parquet_path):
        raise HTTPException(status_code=404, detail="Файл training_data.parquet не найден")
    try:
        df = pd.read_parquet(parquet_path)
        if id_column not in df.columns:
            raise HTTPException(status_code=400, detail=f"Колонка {id_column} не найдена в данных")
        ts_df = df[df[id_column].astype(str) == str(ts_id)]
        if ts_df.empty:
            raise HTTPException(status_code=404, detail="Временной ряд с таким id не найден")
        # Ограничиваем до 50 точек, равномерно по всему ряду
        total = len(ts_df)
        if total > 50:
            idx = (ts_df.index.to_numpy() if hasattr(ts_df.index, 'to_numpy') else ts_df.index.values)
            step = total / 50
            selected_idx = [int(i * step) for i in range(50)]
            ts_df = ts_df.iloc[selected_idx]
        data = ts_df.astype(str).to_dict(orient="records")
        columns = list(ts_df.columns)
        return JSONResponse({"columns": columns, "rows": data, "total_rows": len(ts_df)})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при чтении временного ряда: {str(e)}")

async def analyze_dataframe(df):
    df = df.astype(str)
    columns = list(df.columns)
    rows = df.values.tolist()
    total = len(df)
    total_cells = df.size
    missing_cells = int(
        df.isnull().values.sum() +
        sum((df[col] == '').sum() for col in df.columns if df[col].dtype == object)
    )
    percent = (missing_cells / total_cells) * 100 if total_cells else 0
    date_idx = next((i for i, col in enumerate(columns) if any(x in col.lower() for x in ['date', 'время', 'time'])), -1)
    bin_count = 12
    bin_size = int(np.ceil(total / bin_count)) if total else 1
    bins = []
    for i in range(bin_count):
        start = i * bin_size
        end = min((i + 1) * bin_size, total)
        bin_rows = df.iloc[start:end]
        bin_missing = int(
            bin_rows.isnull().values.sum() +
            sum((bin_rows[col] == '').sum() for col in bin_rows.columns if bin_rows[col].dtype == object)
        )
        name = str(i + 1)
        if date_idx != -1 and not bin_rows.empty:
            first = str(bin_rows.iloc[0, date_idx])[:10]
            last = str(bin_rows.iloc[-1, date_idx])[:10]
            name = f"{first} - {last}"
        bins.append({"name": name, "missing": bin_missing, "total": len(bin_rows)})
    return {
        "columns": columns,
        "rows": rows[:10],
        "total": total,
        "missing": missing_cells,
        "percent": percent,
        "bins": bins
    }

@router.post("/analyze-data")
async def analyze_data(
    file: UploadFile = File(None),
    session_id: str = Form(None)
):
    loop = asyncio.get_event_loop()
    try:
        def load_df():
            if session_id:
                session_dir = os.path.join("training_sessions", session_id)
                parquet_path = os.path.join(session_dir, "original_file.parquet")
                if not os.path.exists(parquet_path):
                    raise HTTPException(status_code=404, detail="Файл training_data.parquet не найден")
                return pd.read_parquet(parquet_path)
            elif file and file.filename:
                filename = file.filename.lower()
                if filename.endswith('.csv'):
                    return pd.read_csv(file.file)
                else:
                    return pd.read_excel(file.file)
            else:
                raise HTTPException(status_code=400, detail="Не передан файл или session_id")
        # Чтение и анализ в отдельном потоке
        with ThreadPoolExecutor() as pool:
            df = await loop.run_in_executor(pool, load_df)
            result = await analyze_dataframe(df)
        return JSONResponse(result)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка анализа данных: {str(e)}")
