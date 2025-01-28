import streamlit as st
import pandas as pd
import yaml
import os
import json
import logging
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from autogluon.timeseries import TimeSeriesPredictor
from pathlib import Path

# Папка и файлы, куда сохраняются модели
MODEL_DIR = "AutogluonModels/TimeSeriesModel"
MODEL_INFO_FILE = "model_info.json"
CONFIG_PATH = "config/config.yaml"


def format_fit_summary(fit_summary) -> str:
    """
    Форматирует fit_summary (словарь) в человекочитаемый текст.
    Можно использовать для отладки или вывода на экран.
    """
    if not fit_summary:
        return "Fit Summary отсутствует."

    summary_str = "### Fit Summary:\n\n"
    if 'total_fit_time' in fit_summary:
        summary_str += f"- **Total Fit Time**: {fit_summary['total_fit_time']:.2f} seconds\n"
    if 'best_model' in fit_summary:
        summary_str += f"- **Best Model**: {fit_summary['best_model']}\n"
    if 'best_model_score' in fit_summary:
        summary_str += f"- **Best Model Score**: {fit_summary['best_model_score']:.4f}\n"

    # Детали по каждой модели, если есть
    if 'model_fit_summary' in fit_summary and fit_summary['model_fit_summary']:
        summary_str += "\n**Model Fit Details:**\n"
        for model_name, model_info in fit_summary['model_fit_summary'].items():
            summary_str += f"\n**Model: {model_name}**\n"
            if isinstance(model_info, dict):
                if 'fit_time' in model_info:
                    summary_str += f"  - Fit Time: {model_info['fit_time']:.2f} seconds\n"
                if 'score' in model_info:
                    summary_str += f"  - Score: {model_info['score']:.4f}\n"
                if 'eval_metric' in model_info:
                    summary_str += f"  - Eval Metric: {model_info['eval_metric']}\n"
            else:
                summary_str += "  - (no details)\n"

    return summary_str


def format_fit_summary_to_df(fit_summary) -> pd.DataFrame:
    """
    Преобразует fit_summary (словарь) в DataFrame для сохранения в Excel.
    """
    if (not fit_summary
            or not isinstance(fit_summary, dict)
            or not fit_summary.get('model_fit_summary')):
        return pd.DataFrame({"Информация": ["Fit Summary пуст или не содержит модельных данных"]})

    data = []
    if 'total_fit_time' in fit_summary:
        data.append({"Метрика": "Total Fit Time", "Значение": f"{fit_summary['total_fit_time']:.2f} seconds"})
    if 'best_model' in fit_summary:
        data.append({"Метрика": "Best Model", "Значение": fit_summary['best_model']})
    if 'best_model_score' in fit_summary:
        data.append({"Метрика": "Best Model Score", "Значение": f"{fit_summary['best_model_score']:.4f}"})

    model_fit = fit_summary.get('model_fit_summary', {})
    for model_name, model_info in model_fit.items():
        data.append({"Метрика": f"Model: {model_name}", "Значение": "---"})
        if isinstance(model_info, dict):
            if 'fit_time' in model_info:
                data.append({"Метрика": f"  Fit Time ({model_name})",
                             "Значение": f"{model_info['fit_time']:.2f} s"})
            if 'score' in model_info:
                data.append({"Метрика": f"  Score ({model_name})",
                             "Значение": f"{model_info['score']:.4f}"})
            if 'eval_metric' in model_info:
                data.append({"Метрика": f"  Eval Metric ({model_name})",
                             "Значение": model_info['eval_metric']})
        else:
            data.append({"Метрика": f"  {model_name} info", "Значение": str(model_info)})

    return pd.DataFrame(data)


def save_model_metadata(dt_col, tgt_col, id_col, static_feats, freq_val,
                       fill_method_val, group_cols_fill_val, use_holidays_val,
                       metric, presets, chosen_models, mean_only):
    """
    Сохраняем метаданные о колонках и настройках модели в JSON-файл (model_info.json).
    Это нужно для восстановления session_state при следующем запуске приложения.
    """
    os.makedirs(MODEL_DIR, exist_ok=True)
    info_dict = {
        "dt_col": dt_col,
        "tgt_col": tgt_col,
        "id_col": id_col,
        "static_feats": static_feats,
        "freq_val": freq_val,
        "fill_method_val": fill_method_val,
        "group_cols_fill_val": group_cols_fill_val,
        "use_holidays_val": use_holidays_val,
        "metric": metric,
        "presets": presets,
        "chosen_models": chosen_models,
        "mean_only": mean_only
    }
    path_json = os.path.join(MODEL_DIR, MODEL_INFO_FILE)
    with open(path_json, "w", encoding="utf-8") as f:
        json.dump(info_dict, f, ensure_ascii=False, indent=2)


def save_results_to_excel(save_path: str) -> bool:
    """
    Сохраняет результаты (predictions, leaderboard, fit_summary и т.д.) в указанный файл.
    - Если расширение .csv: сохраняем предсказания (или df_train) в CSV.
    - Если .xls / .xlsx: сохраняем в Excel с несколькими листами.
    - Убрана запись листа russian_holiday, чтобы не раздувать файл.
    - Если нет данных, выводим warning.
    """
    try:
        df_train = st.session_state.get("df")
        lb = st.session_state.get("leaderboard")
        preds = st.session_state.get("predictions")
        stt_train = st.session_state.get("static_df_train")
        fit_summary_data = st.session_state.get("fit_summary")

        # Проверяем, есть ли вообще что сохранять
        has_data_to_save = any([
            df_train is not None,
            lb is not None,
            preds is not None,
            fit_summary_data,
            (stt_train is not None and not stt_train.empty if stt_train is not None else False),
        ])
        if not has_data_to_save:
            st.warning("Нет данных для сохранения. Сначала загрузите данные и обучите модель.")
            return False

        ext = Path(save_path).suffix.lower()
        if ext == ".csv":
            # Сохраняем в CSV (один лист). Либо preds, либо df_train
            if preds is not None:
                preds_df_to_save = preds.reset_index()
                preds_df_to_save.to_csv(save_path, index=False, encoding="utf-8")
                st.success(f"Предсказания сохранены в CSV: {save_path}")
            else:
                # Если нет preds, сохраним df_train
                if df_train is not None:
                    df_train.to_csv(save_path, index=False, encoding="utf-8")
                    st.success(f"Train-данные сохранены в CSV: {save_path}")
                else:
                    st.error("Нечего сохранять в CSV: нет preds и нет df_train.")
                    return False

        elif ext in (".xlsx", ".xls"):
            # Сохраняем в Excel
            import openpyxl
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                # 1) Predictions
                if preds is not None:
                    pred_df_to_save = preds.reset_index()
                    pred_df_to_save.to_excel(writer, sheet_name="Predictions", index=False)

                # 2) Leaderboard
                if lb is not None:
                    lb.to_excel(writer, sheet_name="Leaderboard", index=False)

                # 3) Fit Summary
                if fit_summary_data:
                    # Можем сохранить "сырые" данные в один лист
                    fs_sheet = pd.DataFrame([{"Fit_Summary": str(fit_summary_data)}])
                    fs_sheet.to_excel(writer, sheet_name="FitSummaryRaw", index=False)

                    # Или подробнее:
                    # fit_sum_df = format_fit_summary_to_df(fit_summary_data)
                    # fit_sum_df.to_excel(writer, sheet_name="FitSummaryDetails", index=False)

                # 4) StaticTrainFeatures (если есть)
                if stt_train is not None and not stt_train.empty:
                    stt_train.to_excel(writer, sheet_name="StaticTrainFeatures", index=False)

                # 5) TrainData
                if df_train is not None:
                    df_train.to_excel(writer, sheet_name="TrainData", index=False)

                # Подсветка лучшей модели в Leaderboard
                if lb is not None and not lb.empty and "Leaderboard" in writer.sheets:
                    sheet_lb = writer.sheets["Leaderboard"]
                    best_idx = lb.iloc[0].name  # Первая строка LB — лучшая модель
                    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    row_excel = best_idx + 2  # +1 для заголовка, +1 ещё
                    for col_idx in range(1, lb.shape[1] + 1):
                        cell = sheet_lb.cell(row=row_excel, column=col_idx)
                        cell.fill = fill_green

            st.success(f"Сохранено в Excel: {save_path}")

        else:
            st.error(f"Неизвестное расширение файла: {ext}. Используйте .csv или .xlsx")
            return False

        return True

    except Exception as ex:
        st.error(f"Ошибка сохранения: {ex}")
        return False


def load_model_metadata():
    """
    Загружаем настройки (dt_col, tgt_col и т.д.) из model_info.json, 
    чтобы при повторном запуске можно было автоматически заполнить session_state.
    """
    path_json = os.path.join(MODEL_DIR, MODEL_INFO_FILE)
    if not os.path.exists(path_json):
        return None
    try:
        with open(path_json, "r", encoding="utf-8") as f:
            info = json.load(f)
        return info
    except Exception:
        return None


def try_load_existing_model():
    """
    Если в папке MODEL_DIR есть ранее обученная модель и model_info.json, 
    то загружаем её и восстанавливаем настройки (дату, ID, target, freq и т.д.)
    в st.session_state.
    """
    if not os.path.exists(MODEL_DIR):
        st.info("Папка с моделью не найдена — модель не загружена.")
        return

    predictor_path = os.path.join(MODEL_DIR, "predictor.pkl")
    if not os.path.exists(predictor_path):
        st.info("Файл predictor.pkl не найден — модель ещё не обучалась.")
        return

    try:
        loaded_predictor = TimeSeriesPredictor.load(MODEL_DIR)
        st.session_state["predictor"] = loaded_predictor
        st.info(f"Загружена ранее обученная модель из {MODEL_DIR}")

        # Загрузим метаданные
        meta = load_model_metadata()
        if meta:
            st.session_state["dt_col_key"] = meta.get("dt_col", "<нет>")
            st.session_state["tgt_col_key"] = meta.get("tgt_col", "<нет>")
            st.session_state["id_col_key"]  = meta.get("id_col", "<нет>")
            st.session_state["static_feats_key"] = meta.get("static_feats", [])
            st.session_state["freq_key"] = meta.get("freq_val", "auto (угадать)")
            st.session_state["fill_method_key"] = meta.get("fill_method_val", "None")
            st.session_state["group_cols_for_fill_key"] = meta.get("group_cols_fill_val", [])
            st.session_state["use_holidays_key"] = meta.get("use_holidays_val", False)
            st.session_state["metric_key"] = meta.get("metric", "MASE (Mean absolute scaled error)")
            st.session_state["presets_key"] = meta.get("presets", "medium_quality")
            st.session_state["models_key"] = meta.get("chosen_models", ["* (все)"])
            st.session_state["mean_only_key"] = meta.get("mean_only", False)

            st.info("Настройки восстановлены из model_info.json (колонки, freq, праздники и т.д.).")
    except Exception as e:
        st.warning(f"Ошибка при загрузке ранее обученной модели: {e}")
